from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def format(file):



    workbook = load_workbook(file)

    # 2. Select a specific sheet (or the active sheet)
    sheet = workbook.active  # or workbook['Sheet1'] if you know the sheet name

    row_count = sheet.max_row
    column_count = sheet.max_column
    sheet.sheet_view.showGridLines = False

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for row_num in range(1, row_count+1):
        for col_num in range(1, column_count+1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment  = Alignment(vertical = 'top')
            cell.border = thin_border
            

    # Set word Wrap to Remarks columns
    word_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for row in range(2, row_count + 1):
        sheet[f'M{row}'].alignment = word_wrap
        sheet[f'Q{row}'].alignment = word_wrap
    


    # format header row and set 
    header_font = Font(bold=True, color='BADC97')
    cells = 'ABCDEFGHIJKLMNOPQRS'
    green_fill = PatternFill(start_color='004000', end_color='004000', fill_type='solid')

    for col in cells:
        sheet[f'{col}1'].font = header_font
        sheet[f'{col}1'].fill = green_fill
        sheet[f'{col}1'].alignment = Alignment(horizontal='center')
        adjusted_width = len(sheet[f'{col}1'].value) * 1.2
        sheet.column_dimensions[col].width = adjusted_width


    sheet.column_dimensions['B'].width += 2
    sheet.column_dimensions['K'].width += 2
    sheet.column_dimensions['Q'].width = 100


    # Save the modified workbook
    workbook.save(file)

